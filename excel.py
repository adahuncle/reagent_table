import pandas as pd
import sqlite3
import json
import os
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image as PILImage, ImageDraw, ImageFont

DB_PATH = "compounds.db"
EXCEL_PATH = "compound_report.xlsx"
TEMPLATE_DIR = "templates"
DEFAULT_TEMPLATE = "Clemson.json"

def load_template(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def summarize_text(text, max_lines=5):
    if not isinstance(text, str):
        return text
    parts = [p.strip(" ;,") for p in text.split(";") if p.strip()]
    return "\n " + "\n ".join(parts[:max_lines]) + ("\n..." if len(parts) > max_lines else "")

def generate_hazards_summary(row):
    hazard_fields = [
        "Other Safety Information",
        "Corrosivity",
        "Fire Fighting",
        "Exposure Control and Personal Protection",
        "First Aid Measures",
        "Accidental Release Measures",
    ]
    summary_parts = []
    for field in hazard_fields:
        val = row.get(field, "")
        if pd.notna(val) and str(val).strip():
            summary_parts.append(f"{field}: {val}")
    summary = "\n".join(summary_parts)
    if len(summary.split()) > 40:
        words = summary.split()
        summary = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))
    return summary

def make_composite_image(image_path, text_lines, text_position="bottom", font_size=14, box_opacity=255, padding=8):
    base_img = PILImage.open(image_path).convert("RGBA")

    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except OSError:
        font = ImageFont.load_default()

    draw = ImageDraw.Draw(base_img)
    spacing = 4

    # Calculate total text height
    line_heights = [
        draw.textbbox((0, 0), line, font=font)[3] - draw.textbbox((0, 0), line, font=font)[1]
        for line in text_lines
    ]
    total_text_height = sum(line_heights) + spacing * (len(text_lines) - 1)
    max_line_width = max(draw.textlength(line, font=font) for line in text_lines)

    # Define box size and position
    box_height = total_text_height + 2 * padding
    box_width = base_img.width  # full width of image

    if text_position == "top":
        box_y = 0
    else:
        box_y = base_img.height - box_height

    # Draw text background box
    box = PILImage.new("RGBA", (box_width, box_height), (255, 255, 255, box_opacity))
    base_img.alpha_composite(box, (0, box_y))

    # Draw text on top of the box
    draw = ImageDraw.Draw(base_img)
    y = box_y + padding
    for line in text_lines:
        draw.text((padding, y), line, font=font, fill="black")
        y += font_size + spacing

    return base_img

def choose_template():
    templates = [f for f in os.listdir(TEMPLATE_DIR) if f.endswith(".json")]
    if not templates:
        raise FileNotFoundError("❌ No templates found in the templates directory.")
    print("\n📁 Available templates:")
    for idx, name in enumerate(templates):
        print(f"[{idx}] {name}")
    choice = input("Enter template number to use: ").strip()
    if not choice.isdigit() or not (0 <= int(choice) < len(templates)):
        raise ValueError("❌ Invalid selection.")
    return os.path.join(TEMPLATE_DIR, templates[int(choice)])

def generate_excel_from_template(template_path=None):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(f"❌ Database not found: {DB_PATH}")
    if not os.path.exists(TEMPLATE_DIR):
        raise FileNotFoundError(f"❌ Template directory missing: {TEMPLATE_DIR}")

    if not template_path:
        template_path = choose_template()

    template = load_template(template_path)

    valid_types = {"text", "image", "composite", "blank", "computed"}
    for col in template["columns"]:
        if col.get("type") not in valid_types:
            raise ValueError(f"❌ Unknown column type: {col.get('type')}")

    conn = sqlite3.connect(DB_PATH)
    df_core = pd.read_sql_query("SELECT * FROM compounds", conn)
    df_props = pd.read_sql_query("SELECT * FROM compound_properties_wide", conn)
    conn.close()

    merged = df_core.merge(df_props, on="cid", how="left")
    if merged.empty:
        raise ValueError("❌ No compounds matched between compounds and compound_properties_wide.")

    # 🔬 TEST BLOCK: Select 5 random rows for testing
    print("🧪 Using a random sample of 5 compounds for test export...")
    if len(merged) >= 5:
        merged = merged.sample(5, random_state=42)
    else:
        print("⚠️ Less than 5 compounds in the database. Exporting all available.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reagent Table"

    headers = [col.get("header", "") for col in template["columns"]]
    ws.append(headers)

    for index, row in merged.iterrows():
        excel_row = []
        for col in template["columns"]:
            col_type = col.get("type")
            if col_type == "composite":
                content = []
                for comp in col.get("components", []):
                    if comp.get("type") == "text":
                        field = comp.get("field")
                        value = row.get(field, "")
                        if pd.notna(value):
                            prefix = comp.get("prefix", "")
                            summarize = comp.get("summarize", False)

                            if "hazard" in field.lower() or "safety" in field.lower():
                                value = summarize_text(value)
                            elif summarize:
                                words = str(value).split()
                                if len(words) > 40:
                                    value = " ".join(words[:30]) + "..."

                            if isinstance(value, str) and len(value.split()) > 15:
                                words = value.split()
                                value = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))

                            content.append(f"{prefix}{value}")
                # Each field's wrapped content starts on its own line
                line_chunks = []
                for line in content:
                    words = str(line).split()
                    if len(words) > 15:
                        wrapped = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))
                    else:
                        wrapped = line
                    line_chunks.append(wrapped)
                excel_row.append("\n".join(line_chunks))


            elif col_type == "image":
                excel_row.append("")
            elif col_type == "blank":
                excel_row.append("")
            elif col_type == "computed":
                func_name = col.get("function")
                if func_name == "generate_hazards_summary":
                    value = generate_hazards_summary(row)
                    excel_row.append(value)
            else:
                value = row.get(col.get("field", ""), "")
                if isinstance(value, str):
                    if col.get("summarize", False):
                        value = summarize_text(value)
                    elif len(value.split()) > 15:
                        words = value.split()
                        value = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))
                excel_row.append(value)
        ws.append(excel_row)

    for i, row in merged.iterrows():
        for j, col in enumerate(template["columns"]):
            if col.get("type") == "composite":
                text_lines = []
                image_path = None
                for comp in col.get("components", []):
                    if comp.get("type") == "text":
                        value = row.get(comp["field"], "")
                        if pd.notna(value):
                            prefix = comp.get("prefix", "")
                            text_lines.append(f"{prefix}{value}")
                    elif comp.get("type") == "image":
                        image_path = row.get(comp["field"])
                if image_path and os.path.exists(image_path):
                    composite_img = make_composite_image(
                        image_path, text_lines, text_position=col.get("text_position", "bottom")
                    )
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                        composite_path = tmp.name
                        composite_img.save(composite_path)
                    img = ExcelImage(composite_path)
                    cell = ws.cell(row=i + 2, column=j + 1)
                    img.anchor = cell.coordinate
                    ws.add_image(img)
                    ws.row_dimensions[i + 2].height = max(ws.row_dimensions[i + 2].height or 0, img.height * 0.75)
                    col_letter = get_column_letter(j + 1)
                    ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 10, img.width / 6)

    for j, col in enumerate(template["columns"]):
        col_letter = get_column_letter(j + 1)
        max_length = 0
        for i in range(1, ws.max_row + 1):
            cell = ws.cell(row=i, column=j + 1)
            text = str(cell.value or "")
            line_count = text.count("\n") + 1 if text else 1

            # Always enable text wrap if multiline or composite
            if "\n" in text or col.get("type") == "composite":
                cell.alignment = Alignment(wrap_text=True, vertical="center")

            # Adjust row height if multiline
            if line_count > 1:
                ws.row_dimensions[i].height = max(ws.row_dimensions[i].height or 15, line_count * 15)

            max_length = max(max_length, len(text))
        ws.column_dimensions[col_letter].width = min(
            max(ws.column_dimensions[col_letter].width or 10, max_length / 1.5 + 2),
            70
        )

    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel generated using template: {template_path}")

if __name__ == "__main__":
    try:
        generate_excel_from_template()
    except Exception as e:
        print(f"❌ An error occurred:\n{e}")
