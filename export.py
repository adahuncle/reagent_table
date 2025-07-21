import sqlite3
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image


def formula_to_subscript(formula):
    subscript_map = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
    return ''.join([c.translate(subscript_map) if c.isdigit() else c for c in formula])

def build_remarks(prop_dict):
    lines = []
    if 'form' in prop_dict:
        lines.append(f"Form - {prop_dict['form']}")
    if 'melting_point' in prop_dict:
        lines.append(f"MP - {prop_dict['melting_point']} °C")
    if 'boiling_point' in prop_dict:
        lines.append(f"BP - {prop_dict['boiling_point']} °C")
    if 'density' in prop_dict:
        lines.append(f"Density - {prop_dict['density']} g/ml")
    return '\n'.join(lines)

def build_hazards(hazard_list):
    hazards = list(hazard_list)
    if len(hazards) < 2:
        if len(hazards) == 1:
            hazards.append("Mild eye irritant")
        else:
            hazards.extend(["Mild respiratory irritant", "Mild eye irritant"])
    return '\n'.join(hazards)

def export_to_excel(compounds_df, properties_df, output_path="lab_reagent_table.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reagents"

    headers = [
        "Chemical Formula",
        "Structure and MW (g/mol)",
        "Purpose",
        "Remarks",
        "Hazards"
    ]
    ws.append(headers)

    for _, compound in compounds_df.iterrows():
        cid = compound["cid"]
        name = compound["name"]
        formula = compound["formula"]
        mw = compound["molecular_weight"]
        image_path = compound["image_path"]
        smiles = compound["smiles"]

        props = properties_df[properties_df["cid"] == cid]
        prop_dict = {row["property_name"]: row["property_value"] for _, row in props.iterrows()}

        # Column 1: Name (Formula)
        chem_formula = f"{name} ({formula_to_subscript(formula)})"

        # Column 2: Image + MW
        struct_text = f"MW: {mw:.2f}"

        # Column 3: Purpose (blank)
        purpose = ""

        # Column 4: Remarks
        remarks = build_remarks(prop_dict)

        # Column 5: Hazards
        hazards = build_hazards(props[props["property_name"] == "hazard"]["property_value"])

        # Add row with placeholders
        row = [chem_formula, struct_text, purpose, remarks, hazards]
        ws.append(row)
        img_cell = f"B{ws.max_row}"

        if os.path.exists(image_path):
            try:
                # Resize image to fit cell
                img = Image.open(image_path)
                img.thumbnail((120, 120))
                temp_path = f"_thumb_{os.path.basename(image_path)}"
                img.save(temp_path)

                xl_img = XLImage(temp_path)
                ws.add_image(xl_img, img_cell)
            except Exception as e:
                print(f"⚠️ Could not add image for CID {cid}: {e}")

        # Format cells
        for col in range(1, 6):
            ws[f"{get_column_letter(col)}{ws.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb.save(output_path)
    print(f"✅ Excel file saved as {output_path}")


def inspect_table(conn, table_name):
    try:
        df = pd.read_sql_query(f"SELECT * FROM {table_name} LIMIT 10", conn)
        print(f"\n📋 Columns in '{table_name}':")
        print(list(df.columns))
        print(f"\n🔍 Preview of '{table_name}' data:")
        print(df)
    except Exception as e:
        print(f"⚠️ Error reading table '{table_name}':", e)


def inspect_and_export(db_path="compounds.db"):
    try:
        conn = sqlite3.connect(db_path)
        inspect_table(conn, "compounds")
        inspect_table(conn, "properties")

        compounds_df = pd.read_sql_query("SELECT * FROM compounds", conn)
        properties_df = pd.read_sql_query("SELECT * FROM properties", conn)
        export_to_excel(compounds_df, properties_df)

    except Exception as e:
        print("⚠️ Database connection error:", e)

    finally:
        conn.close()


if __name__ == "__main__":
    inspect_and_export()