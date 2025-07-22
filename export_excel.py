import pandas as pd
import sqlite3
import json
import os
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image as PILImage, ImageDraw, ImageFont

DB_PATH = "compounds.db"
EXCEL_PATH = "compound_report.xlsx"
TEMPLATE_DIR = "templates"
DEFAULT_TEMPLATE = "Clemson.json"

def load_template(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def summarize_text(text, max_lines=5):
    if not isinstance(text, str):
        return text
    parts = [p.strip(" ;,") for p in text.split(";") if p.strip()]
    return "\n• " + "\n• ".join(parts[:max_lines]) + ("\n..." if len(parts) > max_lines else "")

def generate_hazards_summary(row):
    hazard_fields = [
        "Other Safety Information",
        "Corrosivity",
        "Fire Fighting",
        "Exposure Control and Personal Protection",
        "First Aid Measures",
        "Accidental Release Measures",
    ]
    summary_parts = []
    for field in hazard_fields:
        val = row.get(field, "")
        if pd.notna(val) and str(val).strip():
            summary_parts.append(f"{field}: {val}")
    summary = "\n".join(summary_parts)
    if len(summary.split()) > 40:
        words = summary.split()
        summary = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))
    return summary

def make_composite_image(image_path, text_lines, text_position="bottom", font_size=14):
    base_img = PILImage.open(image_path).convert("RGBA")
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except OSError:
        font = ImageFont.load_default()

    spacing = 5
    draw = ImageDraw.Draw(base_img)
    text_height = sum(
        draw.textbbox((0, 0), line, font=font)[3] - draw.textbbox((0, 0), line, font=font)[1]
        for line in text_lines
    ) + spacing * (len(text_lines) - 1)

    new_height = base_img.height + text_height + 10
    new_img = PILImage.new("RGBA", (base_img.width, new_height), "white")

    if text_position == "top":
        new_img.paste(base_img, (0, text_height + 5))
        y_text = 5
    else:
        new_img.paste(base_img, (0, 0))
        y_text = base_img.height + 5

    draw = ImageDraw.Draw(new_img)
    for line in text_lines:
        draw.text((5, y_text), line, font=font, fill="black")
        y_text += font_size + spacing

    return new_img

def choose_template():
    templates = [f for f in os.listdir(TEMPLATE_DIR) if f.endswith(".json")]
    if not templates:
        raise FileNotFoundError("❌ No templates found in the templates directory.")
    print("\n📁 Available templates:")
    for idx, name in enumerate(templates):
        print(f"[{idx}] {name}")
    choice = input("Enter template number to use: ").strip()
    if not choice.isdigit() or not (0 <= int(choice) < len(templates)):
        raise ValueError("❌ Invalid selection.")
    return os.path.join(TEMPLATE_DIR, templates[int(choice)])

def generate_excel_from_template(template_path=None):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(f"❌ Database not found: {DB_PATH}")
    if not os.path.exists(TEMPLATE_DIR):
        raise FileNotFoundError(f"❌ Template directory missing: {TEMPLATE_DIR}")

    if not template_path:
        template_path = choose_template()

    template = load_template(template_path)

    # Validate template column types
    valid_types = {"text", "image", "composite", "blank", "computed"}
    for col in template["columns"]:
        if col.get("type") not in valid_types:
            raise ValueError(f"❌ Unknown column type: {col.get('type')}")

    conn = sqlite3.connect(DB_PATH)
    df_core = pd.read_sql_query("SELECT * FROM compounds", conn)
    df_props = pd.read_sql_query("SELECT * FROM compound_properties_wide", conn)
    conn.close()

    merged = df_core.merge(df_props, on="cid", how="left")
    if merged.empty:
        raise ValueError("❌ No compounds matched between compounds and compound_properties_wide.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reagent Table"

    headers = [col.get("header", "") for col in template["columns"]]
    ws.append(headers)

    for index, row in merged.iterrows():
        excel_row = []
        for col in template["columns"]:
            col_type = col.get("type")
            if col_type == "composite":
                content = []
                for comp in col.get("components", []):
                    if comp.get("type") == "text":
                        field = comp.get("field")
                        value = row.get(field, "")
                        if pd.notna(value):
                            prefix = comp.get("prefix", "")
                            summarize = comp.get("summarize", False)

                            if "hazard" in field.lower() or "safety" in field.lower():
                                value = summarize_text(value)
                            elif summarize:
                                words = str(value).split()
                                if len(words) > 40:
                                    value = " ".join(words[:30]) + "..."

                            if isinstance(value, str) and len(value.split()) > 15:
                                words = value.split()
                                value = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))

                            content.append(f"{prefix}{value}")
                wrapped_lines = []
                for line in content:
                    words = str(line).split()
                    wrapped = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))
                    wrapped_lines.append(wrapped)
                excel_row.append("\n".join(wrapped_lines))

            elif col_type == "image":
                excel_row.append("")
            elif col_type == "blank":
                excel_row.append("")
            elif col_type == "computed":
                func_name = col.get("function")
                if func_name == "generate_hazards_summary":
                    value = generate_hazards_summary(row)
                    excel_row.append(value)
            else:
                value = row.get(col.get("field", ""), "")
                if isinstance(value, str):
                    if col.get("summarize", False):
                        value = summarize_text(value)
                    elif len(value.split()) > 15:
                        words = value.split()
                        value = "\n".join(" ".join(words[i:i+15]) for i in range(0, len(words), 15))
                excel_row.append(value)
        ws.append(excel_row)

    # Handle composite images
    for i, row in merged.iterrows():
        for j, col in enumerate(template["columns"]):
            if col.get("type") == "composite":
                text_lines = []
                image_path = None
                for comp in col.get("components", []):
                    if comp.get("type") == "text":
                        value = row.get(comp["field"], "")
                        if pd.notna(value):
                            prefix = comp.get("prefix", "")
                            text_lines.append(f"{prefix}{value}")
                    elif comp.get("type") == "image":
                        image_path = row.get(comp["field"])
                if image_path and os.path.exists(image_path):
                    composite_img = make_composite_image(
                        image_path, text_lines, text_position=col.get("text_position", "bottom")
                    )
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                        composite_path = tmp.name
                        composite_img.save(composite_path)
                    img = ExcelImage(composite_path)
                    cell = ws.cell(row=i + 2, column=j + 1)
                    img.anchor = cell.coordinate
                    ws.add_image(img)
                    ws.row_dimensions[i + 2].height = max(ws.row_dimensions[i + 2].height or 0, img.height * 0.75)
                    col_letter = get_column_letter(j + 1)
                    ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 10, img.width / 6)

    for j, col in enumerate(template["columns"]):
        col_letter = get_column_letter(j + 1)
        max_length = 0
        for i in range(1, ws.max_row + 1):
            cell = ws.cell(row=i, column=j + 1)
            text = str(cell.value or "")
            word_count = len(text.split())
            max_length = max(max_length, len(text))
            if word_count > 15:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[col_letter].width = min(
            max(ws.column_dimensions[col_letter].width or 10, max_length / 1.5 + 2),
            70
        )

    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel generated using template: {template_path}")

if __name__ == "__main__":
    try:
        generate_excel_from_template()
    except Exception as e:
        print(f"❌ An error occurred:\n{e}")
